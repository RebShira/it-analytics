import datetime
import os
import shutil
import glob
from openpyxl import load_workbook
from it_analytics import data_interactions as di


REPORT_TEMPLATES = "C:\\\\inetpub\\wwwroot\\Python\\it_analytics\\report_templates\\helpdesk_billing"
REPORT_TARGET = "\\\\10.10.100.220\\it_invoices"

# Dictionary where keys are subcategory_id's and values are Excel row numbers on the template.
SUBCATEGORIES_ROWS = \
{
    1: 15, 2: 29, 3: 30, 4: 31, 5: 32, 6: 33, 7: 34, 8: 35, 9: 36, 10: 37,
    11: 38, 12: 39, 13: 40, 14: 41, 15: 42, 16: 43, 17: 44, 18: 45, 19: 46, 20: 47,
    21: 48, 22: 49, 23: 50, 24: 51, 25: 52, 26: 53, 27: 54, 28: 16, 29: 17, 30: 18,
    31: 19, 32: 20, 33: 21, 34: 22, 35: 23, 36: 24, 37: 25, 38: 26, 39: 27, 40: 28,
    41: 15, 42: 15, 43: 15, 44: 15, 45: 55

}

# In workbook: We are always writing to this column...
WRITE_COLUMN = 10

"""
    steps
    
    1. From date param, extract year and check to see if there is a folder named for
        that year in REPORT_TARGET. If not, mkdir. Add \year to REPORT_TARGET in local
        variable, current_report_target.
    
    2. From date param, extract month and year and create string like: X. MM_YYYY,
        where X is the ordinal of the month. Add to local variable, current_report_target.
        Check whether this directory already exists. If it does, flag the end user.
    
    3. Create, in current_report_target, directories for all AMPs.
    
    4. In each of these directories, copy tamplates from REPORT_TEMPLATES, renaming them with
        month & date & AMP like so: YYYYMM_AMP_(Template name)
        
    5. Run the stored procedure: sp_ProcessMonthlyBilling date param, 'No'
    
    6. Get the newly-created recordset (SQL?)
    
    7. Loop through folders, through templates, populate with data.
    
"""


def process_monthly_billing(drpt):

    # (Initial)
    year = drpt.year
    month = drpt.month
    day = drpt.day

    # Run stored procedure to process the billing:
    report_date = datetime.datetime.strptime(str(month) + "/" + str(day) + "/" + str(year)[2:] + " 00:00:00", "%m/%d/%y 00:00:00")
    sql = "EXEC [dbo].[sp_ProcessMonthlyBilling] '" + str(report_date) + "', 'No'"
    di.crud(sql)

    # Get resultant recordset:
    bill_records = di.return_dataset(ID=10068, Format="Pandas", Data=True, month=month, year=year)

    # Create/overwrite new directory:
    if month < 10:
        month = str("0" + str(month))
    else:
        month = str(month)
    prefix = str(year) + "_" + month
    target_path = REPORT_TARGET + "\\" + prefix + "\\"
    prefix += "_"

    if os.path.exists(target_path):
        shutil.rmtree(target_path)
    os.mkdir(target_path)

    amps = di.return_dataset(ID=10066, Format="pandas", Data=True)
    for index, row in amps.iterrows():
        current_amp = row["dept_no"]
        amp_path = target_path + current_amp + "\\"
        os.mkdir(amp_path)
        files = glob.glob(REPORT_TEMPLATES + "/*")
        for file in files:
            shutil.copy(file, amp_path)
        for filename in os.listdir(amp_path):
            new_name = os.path.join(amp_path, prefix + current_amp + "_" + filename)
            os.rename(os.path.join(amp_path, filename), new_name)

            # We've run the stored proc, attained the dateset, and at this point, have our file.
            # Time to populate them:

            if 'FEE_MGM.xlsx' in filename:
                records_filter = bill_records.loc[(bill_records['subcategory_id'] == 41) &
                                                  (bill_records['dept_no'] == current_amp)]

            elif 'FEE_BOOK.xlsx' in filename:
                records_filter = bill_records.loc[(bill_records['subcategory_id'] == 42) &
                                                  (bill_records['dept_no'] == current_amp)]

            elif 'FEE_ASSET.xlsx' in filename:
                records_filter = bill_records.loc[(bill_records['subcategory_id'] == 43) &
                                                  (bill_records['dept_no'] == current_amp)]

            elif 'LEGAL.xlsx' in filename:
                records_filter = bill_records.loc[(bill_records['subcategory_id'] == 44) &
                                                  (bill_records['dept_no'] == current_amp)]

            else:
                records_filter = bill_records.loc[(bill_records['subcategory_id'] != 41) &
                                                  (bill_records['subcategory_id'] != 42) &
                                                  (bill_records['subcategory_id'] != 43) &
                                                  (bill_records['subcategory_id'] != 44) &
                                                  (bill_records['dept_no'] == current_amp)]

            billing_to_excel(new_name, records_filter, drpt=drpt)


def billing_to_excel(filepath, line_items, drpt):
    wb = load_workbook(filepath)
    ws = wb.active
    ws["B8"] = line_items.iloc[0][1]
    ws["B9"] = line_items.iloc[0][2]
    ws["H4"] = drpt
    ws["H5"] = line_items.iloc[0][3]
    for index, row in line_items.iterrows():
        subcategory_id = row["subcategory_id"]
        amount = row["amount"]
        target_row = SUBCATEGORIES_ROWS.get(subcategory_id)
        ws.cell(row=target_row, column=WRITE_COLUMN).value = amount
    wb.save(filepath)
    wb.close()
    return


def test_openpyxl():
    test_path = REPORT_TEMPLATES + "\\dummy.xlsx"
    wb = load_workbook(test_path)
    ws = wb.active
    ws.cell(row=1, column=1).value = "Andrew will be a good guitarist someday."
    wb.save(test_path)
    wb.close()
    # IT WORKS!!!!!


# test_openpyxl()
process_monthly_billing(datetime.date(2024, 8, 31))